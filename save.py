from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class Save:

    def __init__(self, context):
        self.wb = Workbook()
        now = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.file_name = Path(__file__).parent.joinpath("out", f'{now}.xlsx')
        self._context = context

        self.res_hua_wu = context.res_hua_wu
        self.res_rest = context.res_rest
        self.day_hua_wu = context.rest.day_hua_wu

    def s1(self):
        title = '排休'
        sheet = self.wb.create_sheet(title, index=1)
        sheet['A1'] = '日期'
        for i, day in enumerate(self._context.rest.all_days):
            c = get_column_letter(i + 2)
            sheet[f'{c}1'] = day
            for r in range(self._context.rest.amount):
                value = self.res_rest[r, i]
                if value == 1:
                    data = '休'
                else:
                    data = ''
                sheet[f'{c}{r + 2}'] = data

        _max = len(self._context.rest.all_days) + 2
        cl = get_column_letter(_max)
        data1 = self.res_rest.sum(axis=1)
        for j, v in enumerate(data1):
            sheet[f'{cl}{j + 2}'] = v

        _row_max = self._context.rest.amount + 3
        data2 = self.res_rest.sum(axis=0)
        sheet[f'A{_row_max}'] = '总'
        for k, v in enumerate(data2):
            cl = get_column_letter(2 + k)
            sheet[f'{cl}{_row_max}'] = v

        for gidx, grp in enumerate(self._context.rest.groups):
            sheet[f'A{gidx + 2}'] = grp.name

    def s2(self):
        title = '话务'
        sheet = self.wb.create_sheet(title, index=1)
        row, column = self.res_hua_wu.shape

        real_hw = self.res_hua_wu.sum(axis=0).tolist() # 安排预计的话务
        day_hua_wu = self.day_hua_wu


        for i in range(row):
            for j in range(column):
                sheet.cell(row=i + 2, column=j + 2, value=self.res_hua_wu[i, j])

        for k in range(column):
            a = real_hw[k]
            b = day_hua_wu[k]
            c = round(a/b, 4)
            sheet.cell(row=row + 4, column=k + 2, value=a)
            sheet.cell(row=row + 5, column=k + 2, value=b)
            sheet.cell(row=row + 6, column=k + 2, value=c)
        
        sheet.cell(row=row + 4, column=1, value='排班话务')
        sheet.cell(row=row + 5, column=1, value='预测话务')
        sheet.cell(row=row + 6, column=1, value='排班预计接通率')
            

    def s(self):
        self.s1()
        self.s2()
        self.wb.save(self.file_name)

    @classmethod
    def save(cls, data_contex):
        obj = cls(data_contex)
        obj.s()
